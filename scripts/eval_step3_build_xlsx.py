"""評測步驟 3：把 metrics_results.json 彙整成 6 LLM × 4 RAG 矩陣，輸出 .xlsx"""

import json
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

DATA_DIR = Path("/mnt/c/Users/user/Desktop/驗證/數據資料")
RESULTS_FILE = DATA_DIR / "metrics_results.json"
OUTPUT_XLSX = DATA_DIR / "RAG評測結果.xlsx"

LLMS = [
    ("llama3.1:8b", "Llama3.1-8B"),
    ("deepseek-r1:8b", "DeepSeekR1-8B"),
    ("qwen3:8b", "Qwen3-8B"),
    ("deepseek-r1:32b", "DeepSeekR1-32B"),
    ("qwen3:30b", "Qwen3-30B"),
    ("gpt-oss:20b", "GPTOSS-20B"),
]
STRATEGIES = [
    ("naive_rag", "NaiveRAG"),
    ("advanced_rag", "AdvancedRAG"),
    ("graph_only", "GraphRAG"),
    ("agentic_rag", "AgenticRAG"),
]

METRICS = [
    ("correctness", "Correctness 矩陣", True),
    ("faithfulness", "Faithfulness 矩陣", True),
    ("retrieval_precision", "Retrieval Precision 矩陣", True),
    ("response_time", "Response Time 矩陣（秒）", False),
]


def load_results() -> dict:
    return json.loads(RESULTS_FILE.read_text(encoding="utf-8"))


def build_matrix(results: dict, metric_key: str) -> dict:
    """回傳 {(llm_tag, strategy_code): (mean, sd)}"""
    matrix = {}
    for llm_tag, _ in LLMS:
        for strategy_code, _ in STRATEGIES:
            values = []
            for q_idx in range(15):
                key = f"{q_idx}|{llm_tag}|{strategy_code}"
                if key in results:
                    values.append(results[key][metric_key])
            if values:
                mean = statistics.mean(values)
                sd = statistics.stdev(values) if len(values) > 1 else 0.0
                matrix[(llm_tag, strategy_code)] = (mean, sd, len(values))
            else:
                matrix[(llm_tag, strategy_code)] = (None, None, 0)
    return matrix


def write_sheet(wb: Workbook, title: str, matrix: dict, as_percent: bool):
    ws = wb.create_sheet(title=title[:31])
    ws.cell(1, 1, "RAG 架構 \\ LLM").font = Font(bold=True)
    for col, (_, llm_name) in enumerate(LLMS, start=2):
        c = ws.cell(1, col, llm_name)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row, (strategy_code, rag_name) in enumerate(STRATEGIES, start=2):
        ws.cell(row, 1, rag_name).font = Font(bold=True)
        for col, (llm_tag, _) in enumerate(LLMS, start=2):
            mean, sd, n = matrix[(llm_tag, strategy_code)]
            if mean is None:
                ws.cell(row, col, "N/A")
                continue
            if as_percent:
                text = f"{mean * 100:.2f}%\n({sd * 100:.2f})"
            else:
                text = f"{mean:.2f}\n({sd:.2f})"
            cell = ws.cell(row, col, text)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(LLMS) + 2):
        ws.column_dimensions[chr(64 + col)].width = 18


def write_notapplicable_sheet(wb: Workbook):
    ws = wb.create_sheet(title="MEXA_DALI_SemanticAffinity")
    ws.cell(1, 1, "語義評估指標矩陣摘要：MEXA、DALI 與 Semantic Affinity").font = Font(
        bold=True, size=13
    )
    notes = [
        "",
        "MEXA、DALI：標記為「不適用」",
        "原因：",
        "1. MEXA（Kargaran et al., arXiv:2410.05873）需要「平行雙語語料」+「模型隱藏層向量」，"
        "本評測資料僅有繁體中文單一語言問答，且 Ollama API 無法取得模型內部隱藏層向量。",
        "2. DALI（Ravisankar et al., arXiv:2504.09378）是為「選擇題」設計的指標，需要正確/錯誤"
        "選項的表徵比較，本評測 15 題皆為開放式申論題，無選項結構可套用。",
        "",
        "若強行套用（例如另外編造錯誤選項、或用未經設計驗證的翻譯建立平行語料），算出來的數字"
        "不具方法論效度，因此本次評測不輸出這兩項數字，以避免產出無意義或誤導性的結果。",
        "",
        "Semantic Affinity：另行處理",
        "Gong (arXiv:2601.09732) 的原始方法一樣需要多語言語料，但可用簡化替代方案（翻譯15題標準"
        "答案為英文 + BGE-M3 計算跨語言/單語言擴散比率）產出有意義的近似值，目前尚未執行，"
        "如需要請另行確認執行。",
    ]
    for i, line in enumerate(notes, start=2):
        c = ws.cell(i, 1, line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def main():
    results = load_results()
    print(f"載入 {len(results)} 筆結果")

    wb = Workbook()
    wb.remove(wb.active)

    for metric_key, title, as_percent in METRICS:
        matrix = build_matrix(results, metric_key)
        write_sheet(wb, title, matrix, as_percent)
        print(f"已寫入分頁：{title}")

    write_notapplicable_sheet(wb)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_XLSX))
    print(f"完成，已輸出至 {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
