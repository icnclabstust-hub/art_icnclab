"""評測步驟 6：彙整全部指標（含 MEXA/DALI/Semantic Affinity）成矩陣表格 + 長條圖，
輸出到 C:\\Users\\user\\Desktop\\驗證\\數據資料\\RAG評測結果.xlsx
"""

import json
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.styles import Alignment, Font

DATA_DIR = Path("/mnt/c/Users/user/Desktop/驗證/數據資料")
RESULTS_FILE = DATA_DIR / "metrics_results.json"
SEMANTIC_FILE = DATA_DIR / "semantic_metrics_results.json"
OUTPUT_XLSX = DATA_DIR / "RAG評測結果.xlsx"

LLMS = [
    ("llama3.1:8b", "Llama3.1-8B"),
    ("deepseek-r1:8b", "DeepSeekR1-8B"),
    ("qwen3:8b", "Qwen3-8B"),
    ("deepseek-r1:32b", "DeepSeekR1-32B"),
    ("qwen3:30b", "Qwen3-30B"),
    ("gpt-oss:20b", "GPTOSS-20B"),
]
STRATEGIES = [
    ("naive_rag", "NaiveRAG"),
    ("advanced_rag", "AdvancedRAG"),
    ("graph_only", "GraphRAG"),
    ("agentic_rag", "AgenticRAG"),
]

MAIN_METRICS = [
    ("correctness", "Correctness", True),
    ("faithfulness", "Faithfulness", True),
    ("retrieval_precision", "RetrievalPrecision", True),
    ("response_time", "ResponseTime", False),
]
SEMANTIC_METRICS = [
    ("mexa", "MEXA", True),
    ("dali", "DALI", True),
    ("semantic_affinity", "SemanticAffinity", True),
]


def load_results() -> dict:
    return json.loads(RESULTS_FILE.read_text(encoding="utf-8"))


def load_semantic() -> dict:
    return json.loads(SEMANTIC_FILE.read_text(encoding="utf-8"))


def build_main_matrix(results: dict, metric_key: str) -> dict:
    matrix = {}
    for llm_tag, _ in LLMS:
        for strategy_code, _ in STRATEGIES:
            values = []
            for q_idx in range(15):
                key = f"{q_idx}|{llm_tag}|{strategy_code}"
                if key in results:
                    values.append(results[key][metric_key])
            if values:
                mean = statistics.mean(values)
                sd = statistics.stdev(values) if len(values) > 1 else 0.0
                matrix[(llm_tag, strategy_code)] = (mean, sd)
            else:
                matrix[(llm_tag, strategy_code)] = (0.0, 0.0)
    return matrix


def build_semantic_matrix(semantic: dict, metric_key: str) -> dict:
    matrix = {}
    per_q_field = {
        "mexa": "mexa_dali_per_q",
        "dali": "mexa_dali_per_q",
        "semantic_affinity": "affinity_per_q",
    }[metric_key]
    for llm_tag, _ in LLMS:
        for strategy_code, _ in STRATEGIES:
            combo_key = f"{llm_tag}|{strategy_code}"
            entry = semantic.get(combo_key)
            if not entry:
                matrix[(llm_tag, strategy_code)] = (0.0, 0.0)
                continue
            per_q = entry.get(per_q_field, [])
            mean = entry[metric_key]
            sd = statistics.stdev(per_q) if len(per_q) > 1 else 0.0
            matrix[(llm_tag, strategy_code)] = (mean, sd)
    return matrix


def write_matrix_sheet(wb: Workbook, title: str, matrix: dict, as_percent: bool):
    ws = wb.create_sheet(title=title[:31])
    ws.cell(1, 1, "RAG 架構 \\ LLM").font = Font(bold=True)
    for col, (_, llm_name) in enumerate(LLMS, start=2):
        c = ws.cell(1, col, llm_name)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row, (strategy_code, rag_name) in enumerate(STRATEGIES, start=2):
        ws.cell(row, 1, rag_name).font = Font(bold=True)
        for col, (llm_tag, _) in enumerate(LLMS, start=2):
            mean, sd = matrix[(llm_tag, strategy_code)]
            if as_percent:
                text = f"{mean * 100:.2f}%\n({sd * 100:.2f})"
            else:
                text = f"{mean:.2f}\n({sd:.2f})"
            cell = ws.cell(row, col, text)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 額外附上純數值區（供圖表讀取，不含百分比符號/換行）
    base_row = len(STRATEGIES) + 4
    ws.cell(base_row, 1, "（下方為圖表用純數值資料，勿手動編輯）").font = Font(italic=True, size=9)
    ws.cell(base_row + 1, 1, "RAG 架構")
    for col, (_, llm_name) in enumerate(LLMS, start=2):
        ws.cell(base_row + 1, col, llm_name)
    mean_start_row = base_row + 2
    for row, (strategy_code, rag_name) in enumerate(STRATEGIES, start=mean_start_row):
        ws.cell(row, 1, rag_name)
        for col, (llm_tag, _) in enumerate(LLMS, start=2):
            mean, _ = matrix[(llm_tag, strategy_code)]
            ws.cell(row, col, mean * 100 if as_percent else mean)
    sd_start_row = mean_start_row + len(STRATEGIES) + 1
    ws.cell(sd_start_row - 1, 1, "SD")
    for row, (strategy_code, rag_name) in enumerate(STRATEGIES, start=sd_start_row):
        ws.cell(row, 1, rag_name)
        for col, (llm_tag, _) in enumerate(LLMS, start=2):
            _, sd = matrix[(llm_tag, strategy_code)]
            ws.cell(row, col, sd * 100 if as_percent else sd)

    for col in range(1, len(LLMS) + 2):
        ws.column_dimensions[chr(64 + col)].width = 18

    return ws, mean_start_row, sd_start_row


def add_main_bar_chart(ws, title: str, mean_start_row: int, anchor: str):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"四種 RAG × 六種 LLM {title} 比對圖表"
    chart.y_axis.title = title
    chart.x_axis.title = "LLM"
    chart.height = 9
    chart.width = 22

    cats = Reference(
        ws, min_col=2, max_col=7, min_row=mean_start_row - 1, max_row=mean_start_row - 1
    )
    data = Reference(ws, min_col=1, max_col=7, min_row=mean_start_row, max_row=mean_start_row + 3)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def add_top_combo_chart(
    wb: Workbook, sheet_title: str, matrix: dict, as_percent: bool, top_n: int = 3
):
    """挑出表現最好的 top_n 組合，畫帶誤差棒(SD)的比較圖"""
    combo_list = [
        (llm_tag, strategy_code, mean, sd)
        for (llm_tag, strategy_code), (mean, sd) in matrix.items()
    ]
    combo_list.sort(key=lambda x: -x[2])
    top = combo_list[:top_n]

    ws = wb.create_sheet(title=f"{sheet_title}_高表現組合"[:31])
    llm_disp = dict(LLMS)
    strat_disp = dict(STRATEGIES)
    ws.cell(1, 1, "組合")
    ws.cell(1, 2, "平均值")
    ws.cell(1, 3, "SD")
    for i, (llm_tag, strategy_code, mean, sd) in enumerate(top, start=2):
        label = f"{llm_disp[llm_tag]}+{strat_disp[strategy_code]}"
        ws.cell(i, 1, label)
        ws.cell(i, 2, mean * 100 if as_percent else mean)
        ws.cell(i, 3, sd * 100 if as_percent else sd)
    ws.column_dimensions["A"].width = 30

    chart = BarChart()
    chart.type = "col"
    chart.title = f"{sheet_title} 高表現組合比較圖"
    chart.y_axis.title = sheet_title
    chart.height = 9
    chart.width = 16

    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(top))
    values = Reference(ws, min_col=2, min_row=1, max_row=1 + len(top))
    series = Series(values, title_from_data=True)

    err_ref = Reference(ws, min_col=3, min_row=2, max_row=1 + len(top))
    series.errBars = ErrorBars(
        errBarType="both",
        errValType="cust",
        plus=NumDataSource(NumRef(f=err_ref)),
        minus=NumDataSource(NumRef(f=err_ref)),
    )
    chart.series.append(series)
    chart.set_categories(cats)
    ws.add_chart(chart, "F2")


def write_semantic_summary_sheet(wb: Workbook, semantic: dict):
    ws = wb.create_sheet(title="MEXA_DALI_SA說明")
    ws.cell(1, 1, "語義評估指標矩陣摘要：MEXA、DALI 與 Semantic Affinity").font = Font(
        bold=True, size=13
    )
    notes = [
        "",
        "方法學說明（務必了解此限制再解讀數字）：",
        "原始定義（MEXA: Kargaran et al. arXiv:2410.05873；DALI: Ravisankar et al. "
        "arXiv:2504.09378；Semantic Affinity: Gong arXiv:2601.09732）皆需要模型「內部隱藏層"
        "向量」或「選擇題選項」結構。本系統透過 Ollama API 僅能取得文字輸出與句子級 "
        "embedding（BGE-M3），無法取得模型隱藏層，也沒有選擇題選項結構。",
        "",
        "因此本次計算以「BGE-M3 句子向量」取代「模型隱藏層向量」，並以「同一 LLM+RAG "
        "組合底下的15題」作為配對母體取代「選擇題選項」——每個組合各自建立15x15的中英"
        "（將360份生成答案翻譯為英文）相似度矩陣，計算對角線（同題跨語言配對）是否嚴格"
        "大於非對角線（跨題配對）。",
        "",
        "這是有明確交代取捨的簡化方法，用來確保產出的數字是真實計算而非虛構，但不等同於"
        "原始論文的精確方法，解讀與引用時請註明此差異。",
    ]
    for i, line in enumerate(notes, start=2):
        c = ws.cell(i, 1, line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def main():
    results = load_results()
    semantic = load_semantic()
    print(f"載入 {len(results)} 筆主要指標，{len(semantic)} 組語義指標")

    wb = Workbook()
    wb.remove(wb.active)

    anchors = ["I2", "I2", "I2", "I2"]
    for (metric_key, title, as_percent), anchor in zip(MAIN_METRICS, anchors, strict=True):
        matrix = build_main_matrix(results, metric_key)
        ws, mean_row, _ = write_matrix_sheet(wb, f"{title}矩陣", matrix, as_percent)
        add_main_bar_chart(ws, title, mean_row, anchor)
        add_top_combo_chart(wb, title, matrix, as_percent)
        print(f"已完成：{title}")

    for metric_key, title, as_percent in SEMANTIC_METRICS:
        matrix = build_semantic_matrix(semantic, metric_key)
        ws, mean_row, _ = write_matrix_sheet(wb, f"{title}矩陣", matrix, as_percent)
        add_main_bar_chart(ws, title, mean_row, "I2")
        add_top_combo_chart(wb, title, matrix, as_percent)
        print(f"已完成：{title}")

    write_semantic_summary_sheet(wb, semantic)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_XLSX))
    print(f"完成，已輸出至 {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
